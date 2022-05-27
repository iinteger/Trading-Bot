import requests
from bs4 import BeautifulSoup
import time
import FinanceDataReader as fdr
import openpyxl

wb = openpyxl.load_workbook("stock.xlsx")
data = wb["Sheet1"]
base_url = "https://www.dividend.com/etfs/"


# get_rate
df = fdr.DataReader('USD/KRW', "2022")
rate = df.iloc[-1]["Close"]
data.cell(1,2).value = rate


# stock info
i = 3
while True:
    ticker = data.cell(i, 1).value
    print(ticker)
    if ticker is None:
        break

    url = base_url + str(ticker)
    response = requests.get(url)
    html = response.text
    soup = BeautifulSoup(html, "html.parser")

    # price
    price = soup.find("p", class_="t-text-black t-text-sm t-font-normal t-text-left t-leading-normal").text
    price = int(float(price[1:])*rate)
    data.cell(i, 4).value = price

    # operating fee
    oper_fee = float(soup.find("span", class_="t-font-semibold float-right").text[:-1])
    data.cell(i, 12).value = oper_fee

    # div
    div_rate = float(soup.find_all("p", class_="t-text-black t-text-sm t-font-normal t-text-left t-leading-normal")[6].text[:-1])
    data.cell(i, 13).value = div_rate

    div_cash = float(soup.find("p", class_="t-text-mitre-gray-300 t-text-2xs t-font-medium t-text-left t-leading-normal").text[1:])*rate
    data.cell(i, 14).value = div_cash

    i+=1

wb.save("stock.xlsx")